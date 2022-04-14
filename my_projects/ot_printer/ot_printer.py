#!PYTHON3
import datetime
from os import path
from posixpath import split
from dateutil.relativedelta import relativedelta
from calendar import monthrange
from openpyxl import load_workbook


# date variables 

# TODO - let user pick month
time = datetime.datetime.now()
next_month = ( time + relativedelta(months=1)).strftime("%B").lower()
next_month_days = ( time + relativedelta(months=1)).strftime("%-m %-d %Y")
year = time.strftime("%Y")


file_name = f"{next_month}_{year}_ot_signup_list.txt"


# Todo - find name of day for specific dates in month

def find_weekday(date):
  month, day, year = (int(i) for i in date.split( ' ' ))
  weekday = datetime.date(year, month, day)
  return weekday.strftime("%A")
  

# find number of days in the month

def num_days(date):
    month, day, year = (int(i) for i in date.split( ' ' ))
    year_and_month = year, month
    return monthrange(year_and_month[0],year_and_month[1])[1]

# import file from excel 

wb = load_workbook(filename='4_10s.xlsx', data_only=True)
ws = wb['Overtime']


#print(ws["C30"].value)
# Select which range of cells to use (C30:C41 - AD40:41)(col 3-30, row 30-41)
def day_names():
    days = []
    for col in ws.iter_cols(min_col=3, max_col=33, max_row=1):
      for cell in col:
          days.append(cell.value) 
    return(days)



def ot_headers():
  headers = []
  for row in ws.iter_rows(min_row=30, max_row=39, max_col=1):
    for cell in row:
     headers.append(cell.value)
  return headers * 31
  

def find_num_outages():
  outages = []
  for col in ws.iter_cols(min_col=3, max_col=33, min_row=30, max_row=39):
    for cell in col:
      outages.append(cell.value)
  return outages

def zip_lists(headers, outages):
  ot_list = zip(headers, outages)
  ot_list = [ [header,outages] for (header,outages) in ot_list]
  return ot_list



def create_file(filename):
  open(filename, "w")


def write_to_file(filename):
  x = 0
  i = 0
  data = zip_lists(ot_headers(), find_num_outages())
  split_data = [data[i: i + 10] for i in range(0, len(data), 10)]
  
  print(split_data)
  

  number_of_days = num_days(next_month_days)
  
  day_name =  day_names()
  

  #print(data)
  
  
  with open(filename, "w" ) as working_doc:
      working_doc.write(f"{next_month} {year} Overtime\n\n\n")
      
      
      for day_num in range(number_of_days):
        day_name = day_names()
        working_doc.write(f"{day_name[x]} {day_num + 1}\n\n")
        x += 1
        day = split_data[day_num]
        #print(day)
        for time in range(len(day)):
          time_slot = day[time][0]
          num_outages = int(day[time][1])
          while num_outages < 0:
            working_doc.write(f"{time_slot}________________\n")
            num_outages += 1
            i += 1
            working_doc.write("\n")

          

# Todo - styling  -- Take all data and split into lists 40 items long -- split that list into 2 and 
# generate two columns with the data on one sheet.
      
if not path.exists(file_name):
    write_to_file(file_name)
else:
  write_to_file(file_name)
  print("File Exists!")
